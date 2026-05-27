import argparse
import math
import os
import re
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

from openpyxl import Workbook, load_workbook


DEFAULT_WIRES_FILE = "wires.xlsx"
DEFAULT_NETLIST_FILE = "netlist.txt"
DEFAULT_INST_INFO_FILE = "inst_info.txt"
DEFAULT_STENCIL_FILE = "circuit.vss"
DEFAULT_PLACEMENT_OFFSETS_FILE = "placement_offsets.tsv"

WIRE_COLOR = "RGB(0,0,0)"
WIRE_WEIGHT = "1.1 pt"
DEVICE_TEXT_HEIGHT = 0.12
DEVICE_LABEL_TEXT_HEIGHT = 0.10
DEVICE_LABEL_GAP = 0.03
DEVICE_LABEL_WIDTH = 15.0 / 25.4
DEVICE_LABEL_HEIGHT = 0.16
MOS_BODY_PIN_WIDTH = 6.35 / 25.4
MOS_SOURCE_PIN_HEIGHT = 9.525 / 25.4
PAGE_MARGIN = 0.5
SYMBOL_FIT_MODES = ("uniform", "native", "stretch")
WIRE_ADJUST_MODES = ("none", "snap-endpoints", "adapters")
PIN_ADAPTER_WEIGHT = "1.1 pt"
PIN_ADAPTER_EPS = 1e-4
NODE_SIZE = 2.0 / 25.4
POWER_NET_KEYS = {"VDD", "VSS"}


Point = Tuple[float, float]
BBox = Tuple[float, float, float, float]


@dataclass
class Instance:
    name: str
    cell: str
    xy: Point
    orient: str
    bbox: BBox
    dev_type: str


@dataclass
class WireSegment:
    group_id: int
    seg_id: int
    net: str
    obj_type: str
    layer: str
    purpose: str
    p1: Point
    p2: Point


@dataclass
class NetDevice:
    name: str
    dev_type: str
    pins: Dict[str, str]
    model: str


@dataclass
class PlacementOffsetRule:
    scope: str
    target: str
    dx: float
    dy: float


@dataclass
class PinCandidate:
    net: str
    inst_name: str
    pin: str
    virt_point: Point
    page_point: Point


@dataclass
class WireAdjustment:
    group_id: int
    seg_id: int
    endpoint: str
    net: str
    old_point: Point
    new_point: Point
    inst_name: str
    pin: str
    distance: float


class CoordMap:
    def __init__(
        self,
        bounds: BBox,
        margin: float = PAGE_MARGIN,
        preserve_absolute: bool = False,
        flip_y: bool = False,
        scale: float = 1.0,
    ):
        self.bounds = bounds
        self.margin = margin
        self.preserve_absolute = preserve_absolute
        self.flip_y = flip_y
        self.scale = scale

    def point(self, point: Point) -> Point:
        x, y = point
        min_x, min_y, _max_x, max_y = self.bounds

        if self.preserve_absolute:
            out_x = x
            out_y = y
        else:
            out_x = x - min_x + self.margin
            if self.flip_y:
                out_y = max_y - y + self.margin
            else:
                out_y = y - min_y + self.margin

        return (out_x * self.scale, out_y * self.scale)

    def bbox(self, bbox: BBox) -> BBox:
        x1, y1, x2, y2 = bbox
        corners = [
            self.point((x1, y1)),
            self.point((x1, y2)),
            self.point((x2, y1)),
            self.point((x2, y2)),
        ]
        xs = [p[0] for p in corners]
        ys = [p[1] for p in corners]
        return (min(xs), min(ys), max(xs), max(ys))

    def size(self, value: float) -> float:
        return abs(value) * self.scale

    def inverse_point(self, point: Point) -> Point:
        x, y = point
        min_x, min_y, _max_x, max_y = self.bounds
        raw_x = x / self.scale
        raw_y = y / self.scale

        if self.preserve_absolute:
            return (raw_x, raw_y)

        virt_x = raw_x + min_x - self.margin
        if self.flip_y:
            virt_y = max_y - (raw_y - self.margin)
        else:
            virt_y = raw_y + min_y - self.margin
        return (virt_x, virt_y)


DEVICE_PIN_ORDER = {
    "NMOS": ["D", "G", "S", "B"],
    "PMOS": ["D", "G", "S", "B"],
    "RES": ["R_up", "R_down"],
    "CAP": ["C_up", "C_down"],
}

MASTER_CANDIDATES = {
    "NMOS": ["NMOS", "nmos", "nud18ll_ckt", "n25ll_ckt"],
    "PMOS": ["PMOS", "pmos", "pud18ll_ckt", "p25ll_ckt"],
    "RES": ["R", "RES", "res", "rpposab_2t_ckt"],
    "CAP": ["C", "CAP", "cap"],
    "PIN": ["PIN", "pin", "ipin", "opin", "iopin", "input", "output", "io"],
    "UNKNOWN": ["Unknown", "UNKNOWN"],
}

PIN_HINTS = {
    "NMOS": {
        "D": (0.5, 0.5),
        "G": (-0.5, 0.0),
        "S": (0.5, -0.5),
        "B": (0.5, 0.0),
    },
    "PMOS": {
        "D": (0.5, -0.5),
        "G": (-0.5, 0.0),
        "S": (0.5, 0.5),
        "B": (0.5, 0.0),
    },
    "RES": {
        "R_up": (0.0, 0.5),
        "R_down": (0.0, -0.5),
    },
    "CAP": {
        "C_up": (0.0, 0.5),
        "C_down": (0.0, -0.5),
    },
}


def classify_device(name: str, cell: str = "") -> str:
    upper_name = name.upper()
    lower_cell = cell.lower()

    if upper_name.startswith("NM") or lower_cell.startswith("nud") or "nmos" in lower_cell:
        return "NMOS"
    if upper_name.startswith("PM") or lower_cell.startswith("pud") or "pmos" in lower_cell:
        return "PMOS"
    if upper_name.startswith("R") or lower_cell in {"res", "r"} or "res" in lower_cell:
        return "RES"
    if upper_name.startswith("C") or lower_cell in {"cap", "c"} or "cap" in lower_cell:
        return "CAP"
    if upper_name.startswith("PIN") or lower_cell in {"ipin", "opin", "iopin"}:
        return "PIN"
    return "UNKNOWN"


def parse_instances(path: str) -> Dict[str, Instance]:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        content = f.read()

    instances: Dict[str, Instance] = {}
    for block in re.split(r"\n\s*\n", content.strip()):
        name_m = re.search(r"Name:\s+(\S+)", block)
        cell_m = re.search(r"Cell:\s+(\S+)", block)
        xy_m = re.search(r"XY:\s+\((-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\)", block)
        orient_m = re.search(r"Orient:\s+(\S+)", block)
        bbox_m = re.search(
            r"BBox:\s+\(\((-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\)\s+\((-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\)\)",
            block,
        )
        if not (name_m and cell_m and xy_m and orient_m and bbox_m):
            continue

        name = name_m.group(1)
        cell = cell_m.group(1)
        bbox = tuple(float(bbox_m.group(i)) for i in range(1, 5))  # type: ignore[assignment]
        instances[name] = Instance(
            name=name,
            cell=cell,
            xy=(float(xy_m.group(1)), float(xy_m.group(2))),
            orient=orient_m.group(1),
            bbox=bbox,
            dev_type=classify_device(name, cell),
        )
    return instances


def netlist_instance_name(raw_name: str, instances: Dict[str, Instance]) -> str:
    candidates = [raw_name]
    if len(raw_name) > 1:
        candidates.append(raw_name[1:])
    if raw_name.upper().startswith("X") and len(raw_name) > 1:
        candidates.append(raw_name[1:])

    for candidate in candidates:
        if candidate in instances:
            return candidate
    return candidates[-1]


def parse_netlist(path: str, instances: Dict[str, Instance]) -> List[NetDevice]:
    devices: List[NetDevice] = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line or line.startswith("*") or line.startswith("."):
                continue

            tokens = line.split()
            raw_name = tokens[0]
            name = netlist_instance_name(raw_name, instances)
            dev_type = instances[name].dev_type if name in instances else classify_device(name)

            pin_names = DEVICE_PIN_ORDER.get(dev_type)
            if pin_names is None:
                pin_count = max(2, len(tokens) - 2)
                pin_names = [f"P{i + 1}" for i in range(pin_count)]
            else:
                pin_count = len(pin_names)

            if len(tokens) < 1 + pin_count:
                continue

            pins = dict(zip(pin_names, tokens[1 : 1 + pin_count]))
            model = tokens[1 + pin_count] if len(tokens) > 1 + pin_count else ""
            devices.append(NetDevice(name=name, dev_type=dev_type, pins=pins, model=model))
    return devices


def read_wires_xlsx(path: str) -> List[WireSegment]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise ValueError(f"{path} is empty")

    columns = {str(value).strip().lower(): idx for idx, value in enumerate(header) if value is not None}
    required = ["group_id", "seg_id", "net", "obj_type", "layer", "purpose", "x1", "y1", "x2", "y2"]
    missing = [col for col in required if col not in columns]
    if missing:
        raise ValueError(f"{path} missing required columns: {', '.join(missing)}")

    segments: List[WireSegment] = []
    for excel_row, row in enumerate(rows, start=2):
        if not row or all(value is None for value in row):
            continue
        try:
            segment = WireSegment(
                group_id=int(row[columns["group_id"]]),
                seg_id=int(row[columns["seg_id"]]),
                net=str(row[columns["net"]]),
                obj_type=str(row[columns["obj_type"]]),
                layer=str(row[columns["layer"]]),
                purpose=str(row[columns["purpose"]]),
                p1=(float(row[columns["x1"]]), float(row[columns["y1"]])),
                p2=(float(row[columns["x2"]]), float(row[columns["y2"]])),
            )
        except Exception as exc:
            raise ValueError(f"Invalid wire row {excel_row}: {row}") from exc
        segments.append(segment)

    segments.sort(key=lambda seg: (seg.group_id, seg.seg_id))
    return segments


def parse_placement_offsets(path: str) -> List[PlacementOffsetRule]:
    if not path or not os.path.exists(path):
        return []

    rules: List[PlacementOffsetRule] = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line_no, raw_line in enumerate(f, start=1):
            line = raw_line.split("#", 1)[0].strip()
            if not line:
                continue

            parts = re.split(r"[\t, ]+", line)
            if len(parts) >= 4 and parts[0].lower() in {"scope", "kind"}:
                continue
            if len(parts) != 4:
                raise ValueError(f"Invalid placement offset row {line_no}: {raw_line.rstrip()}")

            scope, target, dx, dy = parts
            scope = scope.lstrip("\ufeff").lower()
            if scope not in {"global", "type", "dev_type", "cell", "inst", "instance", "name"}:
                raise ValueError(f"Invalid placement offset scope at row {line_no}: {scope}")
            rules.append(PlacementOffsetRule(scope=scope, target=target, dx=float(dx), dy=float(dy)))
    return rules


def write_wires_xlsx(path: str, wires: Sequence[WireSegment]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["group_id", "seg_id", "net", "obj_type", "layer", "purpose", "x1", "y1", "x2", "y2"])
    for wire in wires:
        ws.append([
            wire.group_id,
            wire.seg_id,
            wire.net,
            wire.obj_type,
            wire.layer,
            wire.purpose,
            wire.p1[0],
            wire.p1[1],
            wire.p2[0],
            wire.p2[1],
        ])
    wb.save(path)


def write_adjustments_tsv(path: str, adjustments: Sequence[WireAdjustment]) -> None:
    with open(path, "w", encoding="utf-8") as f:
        f.write("group_id\tseg_id\tendpoint\tnet\told_x\told_y\tnew_x\tnew_y\tinst\tpin\tdistance\n")
        for item in adjustments:
            f.write(
                f"{item.group_id}\t{item.seg_id}\t{item.endpoint}\t{item.net}"
                f"\t{item.old_point[0]:.8g}\t{item.old_point[1]:.8g}"
                f"\t{item.new_point[0]:.8g}\t{item.new_point[1]:.8g}"
                f"\t{item.inst_name}\t{item.pin}\t{item.distance:.8g}\n"
            )


def instance_size(inst: Instance) -> Point:
    x1, y1, x2, y2 = inst.bbox
    return (abs(x2 - x1), abs(y2 - y1))


def placement_offset(inst: Instance, rules: Sequence[PlacementOffsetRule] = ()) -> Point:
    dx = 0.0
    dy = 0.0
    for rule in rules:
        target = rule.target
        applies = False
        if rule.scope == "global":
            applies = target == "*" or target.lower() in {"all", "global"}
        elif rule.scope in {"type", "dev_type"}:
            applies = target.upper() == inst.dev_type.upper()
        elif rule.scope == "cell":
            applies = target.lower() == inst.cell.lower()
        elif rule.scope in {"inst", "instance", "name"}:
            applies = target == inst.name

        if applies:
            dx += rule.dx
            dy += rule.dy
    return (dx, dy)


def visio_anchor_to_center_shift(inst: Instance, actual_width: float, actual_height: float) -> Point:
    if inst.dev_type in {"NMOS", "PMOS"}:
        base_dx, base_dy = (actual_width / 2, 0.0)
    elif inst.dev_type in {"RES", "CAP"}:
        base_dx, base_dy = (0.0, -actual_height / 2)
    else:
        return (0.0, 0.0)

    return transform_offset(base_dx, base_dy, inst.orient.upper())


def placement_center(inst: Instance, rules: Sequence[PlacementOffsetRule] = ()) -> Point:
    manual_dx, manual_dy = placement_offset(inst, rules)
    return (inst.xy[0] + manual_dx, inst.xy[1] + manual_dy)


def centered_instance_bbox(inst: Instance, rules: Sequence[PlacementOffsetRule] = ()) -> BBox:
    width, height = instance_size(inst)
    cx, cy = placement_center(inst, rules)
    return (cx - width / 2, cy - height / 2, cx + width / 2, cy + height / 2)


def combined_bounds(
    instances: Iterable[Instance],
    wires: Iterable[WireSegment],
    placement_offsets: Sequence[PlacementOffsetRule] = (),
) -> BBox:
    xs: List[float] = []
    ys: List[float] = []
    for inst in instances:
        x1, y1, x2, y2 = centered_instance_bbox(inst, placement_offsets)
        xs.extend([x1, x2])
        ys.extend([y1, y2])
    for wire in wires:
        xs.extend([wire.p1[0], wire.p2[0]])
        ys.extend([wire.p1[1], wire.p2[1]])

    if not xs or not ys:
        raise ValueError("No drawable geometry found")
    return (min(xs), min(ys), max(xs), max(ys))


def rotated_90(orient: str) -> bool:
    return orient in {"R90", "R270", "MXR90", "MYR90"}


def transform_offset(x: float, y: float, orient: str) -> Point:
    def rotate(px: float, py: float, angle: float) -> Point:
        cos_a, sin_a = math.cos(angle), math.sin(angle)
        return (px * cos_a - py * sin_a, px * sin_a + py * cos_a)

    if orient == "R90":
        return rotate(x, y, math.pi / 2)
    if orient == "R180":
        return rotate(x, y, math.pi)
    if orient == "R270":
        return rotate(x, y, 3 * math.pi / 2)
    if orient == "MX":
        return (x, -y)
    if orient == "MY":
        return (-x, y)
    if orient == "MXR90":
        return rotate(x, -y, math.pi / 2)
    if orient == "MYR90":
        return rotate(-x, y, math.pi / 2)
    return (x, y)


def expected_pin_point(
    inst: Instance,
    pin: str,
    placement_offsets: Sequence[PlacementOffsetRule] = (),
) -> Point:
    hints = PIN_HINTS.get(inst.dev_type, {})
    rx, ry = hints.get(pin, (0.0, 0.0))
    cx, cy = placement_center(inst, placement_offsets)
    width, height = instance_size(inst)
    ox, oy = transform_offset(rx * width, ry * height, inst.orient)
    return (cx + ox, cy + oy)


def expected_mos_body_pin_point(
    inst: Instance,
    placement_offsets: Sequence[PlacementOffsetRule] = (),
) -> Point:
    if inst.dev_type not in {"NMOS", "PMOS"}:
        return expected_pin_point(inst, "B", placement_offsets)

    cx, cy = placement_center(inst, placement_offsets)
    ox, oy = transform_offset(MOS_BODY_PIN_WIDTH, 0.0, inst.orient)
    return (cx + ox, cy + oy)


def expected_mos_source_pin_point(
    inst: Instance,
    placement_offsets: Sequence[PlacementOffsetRule] = (),
) -> Point:
    if inst.dev_type not in {"NMOS", "PMOS"}:
        return expected_pin_point(inst, "S", placement_offsets)

    source_y = MOS_SOURCE_PIN_HEIGHT / 2 if inst.dev_type == "PMOS" else -MOS_SOURCE_PIN_HEIGHT / 2
    cx, cy = placement_center(inst, placement_offsets)
    ox, oy = transform_offset(MOS_BODY_PIN_WIDTH, source_y, inst.orient)
    return (cx + ox, cy + oy)


def mos_body_start_vertices(
    inst: Instance,
    body_net: str,
    vertices_by_net: Dict[str, Set[Point]],
    pin_vertex_threshold: float,
    placement_offsets: Sequence[PlacementOffsetRule],
) -> Set[Tuple[str, Point]]:
    net_key = normalize_net(body_net)
    starts: Set[Tuple[str, Point]] = set()
    for body_point in (
        expected_pin_point(inst, "B", placement_offsets),
        expected_mos_body_pin_point(inst, placement_offsets),
    ):
        body_vertex = nearest_wire_vertex(body_net, vertices_by_net, body_point, pin_vertex_threshold)
        if body_vertex is not None:
            starts.add((net_key, point_key(body_vertex)))
    return starts


def apply_orientation(shape, orient: str) -> None:
    angle_map = {
        "R0": 0.0,
        "R90": math.pi / 2,
        "R180": math.pi,
        "R270": 3 * math.pi / 2,
    }
    if orient in angle_map:
        shape.CellsU("Angle").ResultIU = angle_map[orient]
    elif orient == "MX":
        shape.CellsU("FlipY").FormulaU = "1"
    elif orient == "MY":
        shape.CellsU("FlipX").FormulaU = "1"
    elif orient == "MXR90":
        shape.CellsU("FlipY").FormulaU = "1"
        shape.CellsU("Angle").ResultIU = math.pi / 2
    elif orient == "MYR90":
        shape.CellsU("FlipX").FormulaU = "1"
        shape.CellsU("Angle").ResultIU = math.pi / 2


def build_master_index(stencil) -> Dict[str, object]:
    masters: Dict[str, object] = {}
    if stencil is None:
        return masters

    try:
        for idx in range(1, stencil.Masters.Count + 1):
            master = stencil.Masters.Item(idx)
            for attr in ("NameU", "Name"):
                try:
                    name = getattr(master, attr)
                    if name:
                        masters[str(name).lower()] = master
                except Exception:
                    pass
    except Exception:
        pass
    return masters


def find_master(master_index: Dict[str, object], inst: Instance):
    candidates = [inst.cell, inst.name]
    candidates.extend(MASTER_CANDIDATES.get(inst.dev_type, []))
    candidates.extend(MASTER_CANDIDATES["UNKNOWN"])
    for candidate in candidates:
        master = master_index.get(candidate.lower())
        if master is not None:
            return master
    return None


def draw_fallback_instance(page, inst: Instance, center: Point, width: float, height: float):
    cx, cy = center
    shift_x, shift_y = visio_anchor_to_center_shift(inst, width, height)
    cx += shift_x
    cy += shift_y
    x1 = cx - width / 2
    y1 = cy - height / 2
    x2 = cx + width / 2
    y2 = cy + height / 2
    shape = page.DrawRectangle(x1, y1, x2, y2)
    shape.Text = ""
    shape.CellsU("LineWeight").FormulaU = "0.8 pt"
    shape.CellsU("TxtHeight").ResultIU = DEVICE_TEXT_HEIGHT
    return shape


def draw_device_label(page, inst: Instance, shape) -> object:
    pin_x = float(shape.CellsU("PinX").ResultIU)
    pin_y = float(shape.CellsU("PinY").ResultIU)
    width = float(shape.CellsU("Width").ResultIU)
    height = float(shape.CellsU("Height").ResultIU)

    label_width = DEVICE_LABEL_WIDTH
    x1 = pin_x + width / 2 + DEVICE_LABEL_GAP
    y2 = pin_y + height / 2
    x2 = x1 + label_width
    y1 = y2 - DEVICE_LABEL_HEIGHT

    label = page.DrawRectangle(x1, y1, x2, y2)
    label.Text = inst.name
    label.CellsU("LinePattern").FormulaU = "0"
    label.CellsU("FillPattern").FormulaU = "0"
    label.CellsU("TxtHeight").ResultIU = DEVICE_LABEL_TEXT_HEIGHT
    try:
        label.CellsU("VerticalAlign").FormulaU = "1"
        label.CellsU("Para.HorzAlign").FormulaU = "0"
    except Exception:
        pass
    return label


def resize_symbol(shape, target_w: float, target_h: float, fit_mode: str) -> None:
    if fit_mode == "native":
        return

    if fit_mode == "stretch":
        shape.CellsU("Width").ResultIU = target_w
        shape.CellsU("Height").ResultIU = target_h
        return

    native_w = float(shape.CellsU("Width").ResultIU)
    native_h = float(shape.CellsU("Height").ResultIU)
    if native_w <= 0 or native_h <= 0:
        return

    scale = min(target_w / native_w, target_h / native_h)
    shape.CellsU("Width").ResultIU = native_w * scale
    shape.CellsU("Height").ResultIU = native_h * scale


def draw_instance(
    page,
    master_index: Dict[str, object],
    inst: Instance,
    coord: CoordMap,
    symbol_fit: str,
    placement_offsets: Sequence[PlacementOffsetRule] = (),
    draw_label: bool = True,
):
    cx, cy = coord.point(placement_center(inst, placement_offsets))
    width, height = instance_size(inst)
    bbox_w = coord.size(width)
    bbox_h = coord.size(height)

    master = find_master(master_index, inst)
    if master is None:
        shape = draw_fallback_instance(page, inst, (cx, cy), bbox_w, bbox_h)
        if draw_label:
            draw_device_label(page, inst, shape)
        return shape

    shape = page.Drop(master, cx, cy)
    if rotated_90(inst.orient):
        resize_symbol(shape, bbox_h, bbox_w, symbol_fit)
    else:
        resize_symbol(shape, bbox_w, bbox_h, symbol_fit)
    apply_orientation(shape, inst.orient)
    shape.Text = ""
    shape.CellsU("TxtHeight").ResultIU = DEVICE_TEXT_HEIGHT
    actual_width = float(shape.CellsU("Width").ResultIU)
    actual_height = float(shape.CellsU("Height").ResultIU)
    shift_x, shift_y = visio_anchor_to_center_shift(inst, actual_width, actual_height)
    if shift_x or shift_y:
        shape.CellsU("PinX").ResultIU = float(shape.CellsU("PinX").ResultIU) + shift_x
        shape.CellsU("PinY").ResultIU = float(shape.CellsU("PinY").ResultIU) + shift_y
    if draw_label:
        draw_device_label(page, inst, shape)
    return shape


def delete_shapes(shapes: Dict[str, object]) -> None:
    for shape in shapes.values():
        try:
            shape.Delete()
        except Exception:
            pass
    shapes.clear()


def draw_wire_segment(page, segment: WireSegment, coord: CoordMap):
    x1, y1 = coord.point(segment.p1)
    x2, y2 = coord.point(segment.p2)
    line = page.DrawLine(x1, y1, x2, y2)
    line.CellsU("LineWeight").FormulaU = WIRE_WEIGHT
    line.CellsU("LineColor").FormulaU = WIRE_COLOR
    line.CellsU("LinePattern").FormulaU = "1"
    return line


def incident_direction(center: Point, other: Point) -> Optional[str]:
    cx, cy = center
    ox, oy = other
    if abs(cy - oy) < PIN_ADAPTER_EPS and abs(cx - ox) >= PIN_ADAPTER_EPS:
        return "R" if ox > cx else "L"
    if abs(cx - ox) < PIN_ADAPTER_EPS and abs(cy - oy) >= PIN_ADAPTER_EPS:
        return "U" if oy > cy else "D"
    return None


def is_t_junction_directions(directions: Sequence[str]) -> bool:
    if len(directions) != 3:
        return False
    horizontal = {direction for direction in directions if direction in {"L", "R"}}
    vertical = {direction for direction in directions if direction in {"U", "D"}}
    return (horizontal == {"L", "R"} and len(vertical) == 1) or (
        vertical == {"U", "D"} and len(horizontal) == 1
    )


def t_junction_points(wires: Sequence[WireSegment]) -> List[Point]:
    incident: Dict[Tuple[str, Point], List[str]] = {}
    for wire in wires:
        net_key = normalize_net(wire.net)
        p1_key = point_key(wire.p1)
        p2_key = point_key(wire.p2)

        dir_from_p1 = incident_direction(p1_key, p2_key)
        if dir_from_p1 is not None:
            incident.setdefault((net_key, p1_key), []).append(dir_from_p1)

        dir_from_p2 = incident_direction(p2_key, p1_key)
        if dir_from_p2 is not None:
            incident.setdefault((net_key, p2_key), []).append(dir_from_p2)

    points: Set[Point] = set()
    for (_net, point), directions in incident.items():
        if is_t_junction_directions(directions):
            points.add(point)

    return sorted(points)


def draw_node(page, master_index: Dict[str, object], coord: CoordMap, point: Point):
    x, y = coord.point(point)
    master = master_index.get("node")
    if master is not None:
        shape = page.Drop(master, x, y)
    else:
        half = NODE_SIZE / 2
        shape = page.DrawOval(x - half, y - half, x + half, y + half)

    shape.CellsU("Width").ResultIU = NODE_SIZE
    shape.CellsU("Height").ResultIU = NODE_SIZE
    shape.CellsU("Angle").ResultIU = 0
    return shape


def draw_t_junction_nodes(
    page,
    master_index: Dict[str, object],
    wires: Sequence[WireSegment],
    coord: CoordMap,
) -> int:
    count = 0
    for point in t_junction_points(wires):
        draw_node(page, master_index, coord, point)
        count += 1
    return count


def draw_plain_line(page, p1: Point, p2: Point, weight: str = WIRE_WEIGHT):
    line = page.DrawLine(p1[0], p1[1], p2[0], p2[1])
    line.CellsU("LineWeight").FormulaU = weight
    line.CellsU("LineColor").FormulaU = WIRE_COLOR
    line.CellsU("LinePattern").FormulaU = "1"
    return line


def draw_orthogonal_adapter(page, start: Point, end: Point) -> int:
    sx, sy = start
    ex, ey = end
    if abs(sx - ex) < PIN_ADAPTER_EPS and abs(sy - ey) < PIN_ADAPTER_EPS:
        return 0
    if abs(sx - ex) < PIN_ADAPTER_EPS or abs(sy - ey) < PIN_ADAPTER_EPS:
        draw_plain_line(page, start, end, PIN_ADAPTER_WEIGHT)
        return 1

    if abs(sx - ex) >= abs(sy - ey):
        elbow = (ex, sy)
    else:
        elbow = (sx, ey)

    draw_plain_line(page, start, elbow, PIN_ADAPTER_WEIGHT)
    draw_plain_line(page, elbow, end, PIN_ADAPTER_WEIGHT)
    return 2


def normalize_net(net: str) -> str:
    return net.strip().upper()


def parse_net_set(value: str) -> Set[str]:
    return {net.strip().upper() for net in value.split(",") if net.strip()}


def filter_wires_by_net(wires: Sequence[WireSegment], skip_nets: Set[str]) -> Tuple[List[WireSegment], int]:
    if not skip_nets:
        return list(wires), 0

    kept = [wire for wire in wires if normalize_net(wire.net) not in skip_nets]
    return kept, len(wires) - len(kept)


def mos_body_nets(devices: Sequence[NetDevice]) -> Set[str]:
    nets: Set[str] = set()
    for device in devices:
        if device.dev_type in {"NMOS", "PMOS"}:
            body_net = device.pins.get("B")
            if body_net:
                nets.add(normalize_net(body_net))
    return nets


def build_wire_graph_by_net(
    wires: Sequence[WireSegment],
) -> Tuple[
    Dict[Tuple[int, int], WireSegment],
    Dict[Tuple[str, Point], Set[Tuple[int, int]]],
]:
    wire_lookup: Dict[Tuple[int, int], WireSegment] = {}
    endpoint_to_segments: Dict[Tuple[str, Point], Set[Tuple[int, int]]] = {}

    for wire in wires:
        seg_key = (wire.group_id, wire.seg_id)
        wire_lookup[seg_key] = wire
        net_key = normalize_net(wire.net)
        for endpoint in (point_key(wire.p1), point_key(wire.p2)):
            endpoint_to_segments.setdefault((net_key, endpoint), set()).add(seg_key)

    return wire_lookup, endpoint_to_segments


def is_t_junction_vertex(
    vertex_key: Tuple[str, Point],
    wire_lookup: Dict[Tuple[int, int], WireSegment],
    endpoint_to_segments: Dict[Tuple[str, Point], Set[Tuple[int, int]]],
) -> bool:
    directions: List[str] = []
    _net_key, point = vertex_key
    for seg_key in endpoint_to_segments.get(vertex_key, set()):
        wire = wire_lookup[seg_key]
        p1_key = point_key(wire.p1)
        p2_key = point_key(wire.p2)
        other = p2_key if point == p1_key else p1_key
        direction = incident_direction(point, other)
        if direction is not None:
            directions.append(direction)
    return is_t_junction_directions(directions)


def same_coord(value1: float, value2: float, eps: float = PIN_ADAPTER_EPS) -> bool:
    return abs(value1 - value2) <= eps


def vertical_wire_x(wire: WireSegment) -> Optional[float]:
    p1_key = point_key(wire.p1)
    p2_key = point_key(wire.p2)
    if same_coord(p1_key[0], p2_key[0]):
        return p1_key[0]
    return None


def y_interval_contains_segment(
    outer_y1: float,
    outer_y2: float,
    segment_y1: float,
    segment_y2: float,
    eps: float = PIN_ADAPTER_EPS,
) -> bool:
    outer_low, outer_high = sorted((outer_y1, outer_y2))
    segment_low, segment_high = sorted((segment_y1, segment_y2))
    return segment_low >= outer_low - eps and segment_high <= outer_high + eps


def mos_source_junction_segments_to_restore(
    segments_to_skip: Set[Tuple[int, int]],
    wire_lookup: Dict[Tuple[int, int], WireSegment],
    endpoint_to_segments: Dict[Tuple[str, Point], Set[Tuple[int, int]]],
    devices: Sequence[NetDevice],
    instances: Dict[str, Instance],
    placement_offsets: Sequence[PlacementOffsetRule],
) -> Set[Tuple[int, int]]:
    segments_to_restore: Set[Tuple[int, int]] = set()
    power_junctions = [
        vertex_key
        for vertex_key in endpoint_to_segments
        if vertex_key[0] in POWER_NET_KEYS
        and is_t_junction_vertex(vertex_key, wire_lookup, endpoint_to_segments)
    ]

    for device in devices:
        if device.dev_type not in {"NMOS", "PMOS"}:
            continue
        source_net = device.pins.get("S")
        inst = instances.get(device.name)
        if not source_net or inst is None:
            continue

        net_key = normalize_net(source_net)
        if net_key not in POWER_NET_KEYS:
            continue

        source_point_key = point_key(expected_mos_source_pin_point(inst, placement_offsets))
        same_x_junctions = [
            junction_point
            for junction_net, junction_point in power_junctions
            if junction_net == net_key and same_coord(junction_point[0], source_point_key[0])
        ]
        if not same_x_junctions:
            continue
        if net_key == "VDD":
            directional_junctions = [
                junction_point
                for junction_point in same_x_junctions
                if junction_point[1] >= source_point_key[1] - PIN_ADAPTER_EPS
            ]
            source_junction = max(directional_junctions or same_x_junctions, key=lambda point: point[1])
        else:
            directional_junctions = [
                junction_point
                for junction_point in same_x_junctions
                if junction_point[1] <= source_point_key[1] + PIN_ADAPTER_EPS
            ]
            source_junction = min(directional_junctions or same_x_junctions, key=lambda point: point[1])

        for seg_key in segments_to_skip:
            wire = wire_lookup[seg_key]
            if normalize_net(wire.net) != net_key:
                continue

            wire_x = vertical_wire_x(wire)
            if wire_x is None or not same_coord(wire_x, source_point_key[0]):
                continue

            p1_key = point_key(wire.p1)
            p2_key = point_key(wire.p2)
            if y_interval_contains_segment(
                source_point_key[1],
                source_junction[1],
                p1_key[1],
                p2_key[1],
            ):
                segments_to_restore.add(seg_key)

    return segments_to_restore


def trace_wire_branch_until_first_t_junction(
    start_vertex_key: Tuple[str, Point],
    first_segment_key: Tuple[int, int],
    wire_lookup: Dict[Tuple[int, int], WireSegment],
    endpoint_to_segments: Dict[Tuple[str, Point], Set[Tuple[int, int]]],
) -> Set[Tuple[int, int]]:
    branch: Set[Tuple[int, int]] = set()
    current_vertex_key = start_vertex_key
    current_segment_key = first_segment_key

    while True:
        if current_segment_key in branch:
            break
        branch.add(current_segment_key)

        wire = wire_lookup[current_segment_key]
        p1_key = point_key(wire.p1)
        p2_key = point_key(wire.p2)
        next_point_key = p2_key if current_vertex_key[1] == p1_key else p1_key
        next_vertex_key = (current_vertex_key[0], next_point_key)
        next_segments = endpoint_to_segments.get(next_vertex_key, set())
        if is_t_junction_vertex(next_vertex_key, wire_lookup, endpoint_to_segments):
            break

        next_degree = len(next_segments)
        if next_degree != 2:
            break

        candidates = [seg_key for seg_key in next_segments if seg_key != current_segment_key]
        if len(candidates) != 1:
            break

        current_vertex_key = next_vertex_key
        current_segment_key = candidates[0]

    return branch


def filter_mos_body_wire_branches(
    wires: Sequence[WireSegment],
    devices: Sequence[NetDevice],
    instances: Dict[str, Instance],
    draw_mos_body_wires: bool,
    pin_vertex_threshold: float = 0.8,
    placement_offsets: Sequence[PlacementOffsetRule] = (),
) -> Tuple[List[WireSegment], int, int]:
    if draw_mos_body_wires:
        return list(wires), 0, 0

    wire_lookup, endpoint_to_segments = build_wire_graph_by_net(wires)
    vertices_by_net = wire_vertices_by_net(wires)
    segments_to_skip: Set[Tuple[int, int]] = set()
    matched_body_pin_count = 0

    for device in devices:
        if device.dev_type not in {"NMOS", "PMOS"}:
            continue
        body_net = device.pins.get("B")
        inst = instances.get(device.name)
        if not body_net or inst is None:
            continue

        graph_vertex_keys = mos_body_start_vertices(
            inst,
            body_net,
            vertices_by_net,
            pin_vertex_threshold,
            placement_offsets,
        )
        if not graph_vertex_keys:
            continue

        matched_body_pin_count += 1
        for graph_vertex_key in graph_vertex_keys:
            start_segments = endpoint_to_segments.get(graph_vertex_key, set())
            for start_segment_key in start_segments:
                segments_to_skip.update(
                    trace_wire_branch_until_first_t_junction(
                        graph_vertex_key,
                        start_segment_key,
                        wire_lookup,
                        endpoint_to_segments,
                    )
                )

    segments_to_skip.difference_update(
        mos_source_junction_segments_to_restore(
            segments_to_skip,
            wire_lookup,
            endpoint_to_segments,
            devices,
            instances,
            placement_offsets,
        )
    )

    kept = [wire for wire in wires if (wire.group_id, wire.seg_id) not in segments_to_skip]
    return kept, len(segments_to_skip), matched_body_pin_count


def distance_sq(p1: Point, p2: Point) -> float:
    return (p1[0] - p2[0]) ** 2 + (p1[1] - p2[1]) ** 2


def point_key(point: Point) -> Point:
    return (round(point[0], 6), round(point[1], 6))


def find_dangling_endpoint_refs(
    wires: Sequence[WireSegment],
) -> List[Tuple[WireSegment, str, Point]]:
    refs_by_net_point: Dict[Tuple[str, Point], List[Tuple[WireSegment, str, Point]]] = {}
    for wire in wires:
        net = normalize_net(wire.net)
        for endpoint, point in (("p1", wire.p1), ("p2", wire.p2)):
            refs_by_net_point.setdefault((net, point_key(point)), []).append((wire, endpoint, point))

    dangling_refs: List[Tuple[WireSegment, str, Point]] = []
    for refs in refs_by_net_point.values():
        if len(refs) == 1:
            dangling_refs.append(refs[0])
    return dangling_refs


def build_pin_candidates(
    page,
    devices: Sequence[NetDevice],
    instances: Dict[str, Instance],
    shapes: Dict[str, object],
    excluded_pins: Set[str],
    coord: CoordMap,
    placement_offsets: Sequence[PlacementOffsetRule] = (),
) -> Dict[str, List[PinCandidate]]:
    candidates: Dict[str, List[PinCandidate]] = {}

    for device in devices:
        inst = instances.get(device.name)
        shape = shapes.get(device.name)
        if inst is None or shape is None:
            continue

        pin_order = DEVICE_PIN_ORDER.get(inst.dev_type, [])
        for pin, net in device.pins.items():
            if pin.upper() in excluded_pins or pin not in pin_order:
                continue

            conn_idx = pin_order.index(pin) + 1
            page_point = connection_page_point(page, shape, conn_idx)
            if page_point is None:
                continue

            candidates.setdefault(normalize_net(net), []).append(
                PinCandidate(
                    net=net,
                    inst_name=device.name,
                    pin=pin,
                    virt_point=coord.inverse_point(page_point),
                    page_point=page_point,
                )
            )

    return candidates


def snap_dangling_wire_endpoints(
    page,
    wires: Sequence[WireSegment],
    devices: Sequence[NetDevice],
    instances: Dict[str, Instance],
    shapes: Dict[str, object],
    coord: CoordMap,
    threshold: float,
    excluded_pins: Set[str],
    placement_offsets: Sequence[PlacementOffsetRule] = (),
) -> List[WireAdjustment]:
    candidates_by_net = build_pin_candidates(page, devices, instances, shapes, excluded_pins, coord, placement_offsets)
    adjustments: List[WireAdjustment] = []
    used_pins: Set[Tuple[str, str]] = set()

    for wire, endpoint, old_virt_point in find_dangling_endpoint_refs(wires):
        net_key = normalize_net(wire.net)
        candidates = candidates_by_net.get(net_key, [])
        if not candidates:
            continue

        available = [
            candidate
            for candidate in candidates
            if (candidate.inst_name, candidate.pin) not in used_pins
        ]
        if not available:
            available = candidates

        best = min(available, key=lambda candidate: distance_sq(old_virt_point, candidate.virt_point))
        distance = math.sqrt(distance_sq(old_virt_point, best.virt_point))
        if distance > threshold:
            continue

        new_virt_point = coord.inverse_point(best.page_point)
        if math.sqrt(distance_sq(old_virt_point, new_virt_point)) <= PIN_ADAPTER_EPS:
            continue

        if endpoint == "p1":
            wire.p1 = new_virt_point
        else:
            wire.p2 = new_virt_point

        used_pins.add((best.inst_name, best.pin))
        adjustments.append(
            WireAdjustment(
                group_id=wire.group_id,
                seg_id=wire.seg_id,
                endpoint=endpoint,
                net=wire.net,
                old_point=old_virt_point,
                new_point=new_virt_point,
                inst_name=best.inst_name,
                pin=best.pin,
                distance=distance,
            )
        )

    return adjustments


def wire_vertices_by_net(wires: Sequence[WireSegment]) -> Dict[str, Set[Point]]:
    vertices: Dict[str, Set[Point]] = {}
    for wire in wires:
        key = normalize_net(wire.net)
        vertices.setdefault(key, set()).add(wire.p1)
        vertices.setdefault(key, set()).add(wire.p2)
    return vertices


def nearest_wire_vertex(
    net: str,
    vertices_by_net: Dict[str, Set[Point]],
    reference_point: Point,
    threshold: float,
) -> Optional[Point]:
    vertices = vertices_by_net.get(normalize_net(net), set())
    if not vertices:
        return None

    best = min(vertices, key=lambda point: distance_sq(point, reference_point))
    if math.sqrt(distance_sq(best, reference_point)) > threshold:
        return None
    return best


def connection_page_point(page, shape, conn_idx: int) -> Optional[Point]:
    try:
        conn_x = shape.CellsU(f"Connections.X{conn_idx}")
        conn_y = shape.CellsU(f"Connections.Y{conn_idx}")
    except Exception:
        return None

    probe = None
    try:
        probe = page.DrawLine(0, 0, 0.01, 0.01)
        probe.CellsU("BeginX").GlueTo(conn_x)
        probe.CellsU("BeginY").GlueTo(conn_y)
        return (
            float(probe.CellsU("BeginX").ResultIU),
            float(probe.CellsU("BeginY").ResultIU),
        )
    except Exception:
        return None
    finally:
        if probe is not None:
            try:
                probe.Delete()
            except Exception:
                pass


def draw_pin_adapters(
    page,
    devices: Sequence[NetDevice],
    instances: Dict[str, Instance],
    shapes: Dict[str, object],
    wires: Sequence[WireSegment],
    coord: CoordMap,
    threshold: float,
    excluded_pins: Set[str],
    placement_offsets: Sequence[PlacementOffsetRule] = (),
) -> int:
    vertices = wire_vertices_by_net(wires)
    adapter_count = 0

    for device in devices:
        inst = instances.get(device.name)
        shape = shapes.get(device.name)
        if inst is None or shape is None:
            continue

        pin_order = DEVICE_PIN_ORDER.get(inst.dev_type, [])
        for pin, net in device.pins.items():
            if pin.upper() in excluded_pins:
                continue
            if pin not in pin_order:
                continue

            conn_idx = pin_order.index(pin) + 1
            visio_point = connection_page_point(page, shape, conn_idx)
            if visio_point is None:
                continue

            pin_virt_point = coord.inverse_point(visio_point)
            virt_point = nearest_wire_vertex(net, vertices, pin_virt_point, threshold)
            if virt_point is None:
                continue

            target_point = coord.point(virt_point)
            if math.sqrt(distance_sq(visio_point, target_point)) <= PIN_ADAPTER_EPS:
                continue

            adapter_count += draw_orthogonal_adapter(page, visio_point, target_point)

    return adapter_count


def validate_inputs(
    instances: Dict[str, Instance],
    devices: Sequence[NetDevice],
    wires: Sequence[WireSegment],
) -> List[str]:
    messages: List[str] = []
    if not instances:
        messages.append("ERROR: no instances parsed from inst_info.txt")
    if not wires:
        messages.append("ERROR: no wire segments parsed from wires.xlsx")

    missing_instances = [dev.name for dev in devices if dev.name not in instances]
    if missing_instances:
        messages.append("WARNING: netlist devices missing in inst_info: " + ", ".join(missing_instances))

    netlist_nets = {normalize_net(net) for dev in devices for net in dev.pins.values()}
    wire_nets = {normalize_net(seg.net) for seg in wires}

    missing_wire_nets = sorted(netlist_nets - wire_nets)
    extra_wire_nets = sorted(wire_nets - netlist_nets)
    if missing_wire_nets:
        messages.append("WARNING: nets in netlist but not in wires.xlsx: " + ", ".join(missing_wire_nets))
    if extra_wire_nets:
        messages.append("WARNING: nets in wires.xlsx but not in parsed netlist pins: " + ", ".join(extra_wire_nets))

    messages.append(f"Parsed {len(instances)} instances, {len(devices)} netlist devices, {len(wires)} wire segments")
    return messages


def draw_visio(
    instances: Dict[str, Instance],
    devices: Sequence[NetDevice],
    wires: Sequence[WireSegment],
    stencil_path: str,
    coord: CoordMap,
    symbol_fit: str,
    wire_adjust: str,
    pin_snap_threshold: float,
    excluded_pins: Set[str],
    placement_offsets: Sequence[PlacementOffsetRule],
    adjusted_wires_path: str,
    adjustment_report_path: str,
    draw_nodes: bool = True,
    visible: bool = True,
) -> None:
    try:
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("pywin32 is required for Visio drawing. Install with: pip install pywin32") from exc

    visio = win32com.client.Dispatch("Visio.Application")
    visio.Visible = visible
    visio.Documents.Add("")
    page = visio.ActivePage

    min_x, min_y, max_x, max_y = coord.bbox(coord.bounds)
    page_sheet = page.PageSheet
    page_sheet.CellsU("PageWidth").ResultIU = max(1.0, max_x - min_x + PAGE_MARGIN)
    page_sheet.CellsU("PageHeight").ResultIU = max(1.0, max_y - min_y + PAGE_MARGIN)

    stencil = None
    if stencil_path and os.path.exists(stencil_path):
        stencil = visio.Documents.OpenEx(os.path.abspath(stencil_path), 64)
    master_index = build_master_index(stencil)

    # Temporary symbols are only used to inspect pin geometry. They are removed
    # before final drawing so the visible order is wires first, devices last.
    temp_shapes: Dict[str, object] = {}
    if wire_adjust in {"snap-endpoints", "adapters"}:
        for inst in instances.values():
            temp_shapes[inst.name] = draw_instance(
                page,
                master_index,
                inst,
                coord,
                symbol_fit,
                placement_offsets,
                draw_label=False,
            )

    adjustments: List[WireAdjustment] = []
    dangling_count = 0
    if wire_adjust == "snap-endpoints":
        dangling_count = len(find_dangling_endpoint_refs(wires))
        adjustments = snap_dangling_wire_endpoints(
            page,
            wires,
            devices,
            instances,
            temp_shapes,
            coord,
            pin_snap_threshold,
            excluded_pins,
            placement_offsets,
        )
        write_wires_xlsx(adjusted_wires_path, wires)
        write_adjustments_tsv(adjustment_report_path, adjustments)

    if wire_adjust != "adapters":
        delete_shapes(temp_shapes)

    for segment in wires:
        draw_wire_segment(page, segment, coord)
    node_count = draw_t_junction_nodes(page, master_index, wires, coord) if draw_nodes else 0

    adapter_count = 0
    if wire_adjust == "adapters":
        adapter_count = draw_pin_adapters(
            page,
            devices,
            instances,
            temp_shapes,
            wires,
            coord,
            pin_snap_threshold,
            excluded_pins,
            placement_offsets,
        )
        delete_shapes(temp_shapes)

    shapes: Dict[str, object] = {}
    for inst in instances.values():
        shapes[inst.name] = draw_instance(page, master_index, inst, coord, symbol_fit, placement_offsets)

    print(f"Drew {len(instances)} instances and {len(wires)} exact Virtuoso wire segments in Visio")
    print(f"Added {node_count} T-junction node(s)" if draw_nodes else "Skipped T-junction nodes")
    if wire_adjust == "snap-endpoints":
        print(f"Found {dangling_count} dangling wire endpoints; snapped {len(adjustments)}")
        print(f"Wrote adjusted wires to {adjusted_wires_path}")
        print(f"Wrote adjustment report to {adjustment_report_path}")
    if wire_adjust == "adapters":
        print(f"Added {adapter_count} local pin adapter line segments")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Draw a Visio schematic from exact Virtuoso wire coordinates."
    )
    parser.add_argument("--wires", default=DEFAULT_WIRES_FILE, help="Path to wires.xlsx")
    parser.add_argument("--netlist", default=DEFAULT_NETLIST_FILE, help="Path to CDL netlist")
    parser.add_argument("--inst-info", default=DEFAULT_INST_INFO_FILE, help="Path to inst_info.txt")
    parser.add_argument("--stencil", default=DEFAULT_STENCIL_FILE, help="Path to circuit.vss")
    parser.add_argument("--scale", type=float, default=1.0, help="Global drawing scale")
    parser.add_argument(
        "--symbol-fit",
        choices=SYMBOL_FIT_MODES,
        default="native",
        help="How to size Visio stencil symbols: uniform preserves aspect ratio, native keeps stencil size, stretch fills Virtuoso BBox",
    )
    parser.add_argument(
        "--wire-adjust",
        choices=WIRE_ADJUST_MODES,
        default="none",
        help="How to compensate for native Visio symbol pin offsets before drawing wires; use snap-endpoints only when endpoint micro-adjustment is desired",
    )
    parser.add_argument(
        "--pin-snap-threshold",
        type=float,
        default=0.8,
        help="Maximum Virtuoso distance used to match a device pin to a same-net wire node",
    )
    parser.add_argument(
        "--exclude-pins",
        default="",
        help="Comma-separated device pins excluded from wire endpoint snapping",
    )
    parser.add_argument(
        "--placement-offsets",
        default=DEFAULT_PLACEMENT_OFFSETS_FILE,
        help="Optional TSV/CSV placement calibration file: scope target dx dy",
    )
    parser.add_argument(
        "--skip-nets",
        default="",
        help="Comma-separated net names to skip before endpoint snapping and drawing, e.g. vdd,vss",
    )
    parser.add_argument(
        "--skip-mos-body-nets",
        action="store_true",
        help="Skip every net connected to an NMOS/PMOS B pin in the CDL netlist",
    )
    parser.add_argument(
        "--draw-mos-b-wires",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="Whether to draw MOS B-pin wire branches. Use --no-draw-mos-b-wires to hide wires from B up to the first T junction.",
    )
    parser.add_argument(
        "--draw-nodes",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Whether to draw visible T-junction node dots. Use --no-draw-nodes to hide node dots.",
    )
    parser.add_argument(
        "--adjusted-wires-output",
        default="wires_adjusted.xlsx",
        help="Output workbook containing wire coordinates after endpoint snapping",
    )
    parser.add_argument(
        "--adjustment-report",
        default="wire_adjustments.tsv",
        help="TSV report listing every snapped wire endpoint",
    )
    parser.add_argument("--preserve-absolute", action="store_true", help="Do not translate coordinates")
    parser.add_argument("--flip-y", action="store_true", help="Flip Y during page coordinate mapping")
    parser.add_argument("--dry-run", action="store_true", help="Validate inputs without opening Visio")
    parser.add_argument("--hidden", action="store_true", help="Start Visio hidden")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    instances = parse_instances(args.inst_info)
    devices = parse_netlist(args.netlist, instances)
    wires = read_wires_xlsx(args.wires)
    placement_offsets = parse_placement_offsets(args.placement_offsets)

    for message in validate_inputs(instances, devices, wires):
        print(message)

    skip_nets = parse_net_set(args.skip_nets)
    if args.skip_mos_body_nets:
        body_nets = mos_body_nets(devices)
        skip_nets.update(body_nets)
        print(f"MOS body nets from netlist: {', '.join(sorted(body_nets)) if body_nets else '(none)'}")

    wires, skipped_wire_count = filter_wires_by_net(wires, skip_nets)
    if skip_nets:
        print(f"Skipping nets: {', '.join(sorted(skip_nets))}")
        print(f"Skipped {skipped_wire_count} wire segments; {len(wires)} remain")

    wires, skipped_body_wire_count, matched_body_pin_count = filter_mos_body_wire_branches(
        wires,
        devices,
        instances,
        draw_mos_body_wires=args.draw_mos_b_wires,
        pin_vertex_threshold=args.pin_snap_threshold,
        placement_offsets=placement_offsets,
    )
    if not args.draw_mos_b_wires:
        print(
            "MOS B wire branch filter: "
            f"matched {matched_body_pin_count} body pin(s), "
            f"skipped {skipped_body_wire_count} wire segment(s), "
            f"{len(wires)} remain"
        )

    bounds = combined_bounds(instances.values(), wires, placement_offsets)
    coord = CoordMap(
        bounds=bounds,
        preserve_absolute=args.preserve_absolute,
        flip_y=args.flip_y,
        scale=args.scale,
    )

    print(f"Virtuoso bounds: x={bounds[0]}..{bounds[2]}, y={bounds[1]}..{bounds[3]}")
    if placement_offsets:
        print(f"Loaded {len(placement_offsets)} placement offset rule(s) from {args.placement_offsets}")
    if args.dry_run:
        print("Dry run complete; Visio was not opened")
        return

    excluded_pins = {pin.strip().upper() for pin in args.exclude_pins.split(",") if pin.strip()}
    if not args.draw_mos_b_wires:
        excluded_pins.add("B")
    draw_visio(
        instances,
        devices,
        wires,
        args.stencil,
        coord,
        symbol_fit=args.symbol_fit,
        wire_adjust=args.wire_adjust,
        pin_snap_threshold=args.pin_snap_threshold,
        excluded_pins=excluded_pins,
        placement_offsets=placement_offsets,
        adjusted_wires_path=args.adjusted_wires_output,
        adjustment_report_path=args.adjustment_report,
        draw_nodes=args.draw_nodes,
        visible=not args.hidden,
    )


if __name__ == "__main__":
    main()
